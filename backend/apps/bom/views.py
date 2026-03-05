import io
import csv
import openpyxl
from decimal import Decimal, InvalidOperation
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, JSONParser

from .models import BOM, BOMItem
from .serializers import BOMSerializer, BOMListSerializer
from apps.master.models import RawMaterial, ProductModel


class BOMListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        boms = BOM.objects.select_related('product_model').all()
        product_id = request.query_params.get('product_model')
        if product_id:
            boms = boms.filter(product_model_id=product_id)
        return Response(BOMListSerializer(boms, many=True).data)

    def post(self, request):
        if request.user.role not in ['ADMIN', 'PRODUCTION_MANAGER', 'STORE_MANAGER']:
            return Response({'error': 'Permission denied'}, status=403)

        product_model_id = request.data.get('product_model')
        if not product_model_id:
            return Response({'error': 'product_model is required'}, status=400)

        # Deactivate existing BOM for this product model
        BOM.objects.filter(product_model_id=product_model_id, is_active=True).update(is_active=False)

        serializer = BOMSerializer(data=request.data)
        if serializer.is_valid():
            bom = serializer.save(created_by=request.user)
            return Response(BOMSerializer(bom).data, status=201)
        return Response(serializer.errors, status=400)


class BOMDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def get_object(self, pk):
        try:
            return BOM.objects.prefetch_related('items__raw_material').get(pk=pk)
        except BOM.DoesNotExist:
            return None

    def get(self, request, pk):
        bom = self.get_object(pk)
        if not bom:
            return Response({'error': 'BOM not found'}, status=404)
        return Response(BOMSerializer(bom).data)

    def put(self, request, pk):
        if request.user.role not in ['ADMIN', 'PRODUCTION_MANAGER', 'STORE_MANAGER']:
            return Response({'error': 'Permission denied'}, status=403)
        bom = self.get_object(pk)
        if not bom:
            return Response({'error': 'BOM not found'}, status=404)

        # Update BOM items
        items_data = request.data.get('items', [])
        bom.notes = request.data.get('notes', bom.notes)
        bom.save()

        # Replace all items
        bom.items.all().delete()
        errors = []
        for item_data in items_data:
            try:
                BOMItem.objects.create(
                    bom=bom,
                    raw_material_id=item_data['raw_material'],
                    qty_per_unit=Decimal(str(item_data['qty_per_unit'])),
                    scrap_percent=Decimal(str(item_data.get('scrap_percent', 0))),
                    process_stage=item_data.get('process_stage', '')
                )
            except Exception as e:
                errors.append(str(e))

        if errors:
            return Response({'error': errors}, status=400)
        return Response(BOMSerializer(self.get_object(pk)).data)

    def delete(self, request, pk):
        if request.user.role != 'ADMIN':
            return Response({'error': 'Permission denied'}, status=403)
        bom = self.get_object(pk)
        if not bom:
            return Response({'error': 'BOM not found'}, status=404)
        # Check if production orders exist
        if bom.product_model.production_orders.exists():
            return Response({'error': 'Cannot delete — production orders exist for this model'}, status=400)
        bom.delete()
        return Response(status=204)


class BOMImportView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def post(self, request):
        if request.user.role not in ['ADMIN', 'PRODUCTION_MANAGER', 'STORE_MANAGER']:
            return Response({'error': 'Permission denied'}, status=403)

        product_model_id = request.data.get('product_model')
        if not product_model_id:
            return Response({'error': 'product_model is required'}, status=400)

        try:
            product_model = ProductModel.objects.get(pk=product_model_id)
        except ProductModel.DoesNotExist:
            return Response({'error': 'Product model not found'}, status=404)

        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'No file provided'}, status=400)

        filename = file_obj.name.lower()
        rows = []

        try:
            if filename.endswith('.csv'):
                content = file_obj.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(content))
                rows = list(reader)
            elif filename.endswith(('.xlsx', '.xls')):
                wb = openpyxl.load_workbook(file_obj)
                ws = wb.active
                headers = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows())]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(v for v in row):
                        rows.append(dict(zip(headers, [str(v).strip() if v is not None else '' for v in row])))
            else:
                return Response({'error': 'Only .csv, .xlsx, .xls files supported'}, status=400)
        except Exception as e:
            return Response({'error': f'File parse error: {str(e)}'}, status=400)

        # Deactivate existing BOM
        BOM.objects.filter(product_model=product_model, is_active=True).update(is_active=False)
        bom = BOM.objects.create(product_model=product_model, created_by=request.user)

        imported, skipped = 0, []
        for i, row in enumerate(rows, start=2):
            item_id = row.get('item_id', '').strip()
            item_name = row.get('item_name', '').strip()

            material = None
            if item_id:
                material = RawMaterial.objects.filter(item_id=item_id).first()
            if not material and item_name:
                material = RawMaterial.objects.filter(item_name__iexact=item_name).first()
            if not material:
                skipped.append(f"Row {i}: item '{item_id or item_name}' not found")
                continue

            try:
                qty = Decimal(str(row.get('qty_per_unit', row.get('quantity', '0'))))
                scrap = Decimal(str(row.get('scrap_percent', row.get('scrap', '0'))))
                stage = row.get('process_stage', row.get('stage', ''))
                BOMItem.objects.create(
                    bom=bom, raw_material=material,
                    qty_per_unit=qty, scrap_percent=scrap, process_stage=stage
                )
                imported += 1
            except (InvalidOperation, Exception) as e:
                skipped.append(f"Row {i}: {str(e)}")

        return Response({
            'bom_id': bom.id,
            'product_model': product_model.model_name,
            'imported': imported,
            'skipped': skipped,
            'total_cost': float(bom.total_cost)
        })


class BOMExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        try:
            bom = BOM.objects.prefetch_related('items__raw_material').get(pk=pk)
        except BOM.DoesNotExist:
            return Response({'error': 'BOM not found'}, status=404)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BOM"
        headers = ['item_id', 'item_name', 'unit', 'qty_per_unit', 'scrap_percent', 'process_stage', 'line_cost']
        ws.append(headers)
        for item in bom.items.select_related('raw_material').all():
            ws.append([
                item.raw_material.item_id,
                item.raw_material.item_name,
                item.raw_material.unit,
                float(item.qty_per_unit),
                float(item.scrap_percent),
                item.process_stage,
                float(item.line_cost)
            ])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="BOM_{bom.product_model.model_id}.xlsx"'
        return response


class BOMSampleDownloadView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BOM Template"
        headers = ['item_id', 'item_name', 'qty_per_unit', 'scrap_percent', 'process_stage']
        ws.append(headers)
        # Sample rows
        ws.append(['RM001', 'Motor Winding Wire', '1.5', '2.0', 'Winding'])
        ws.append(['RM002', 'Pump Housing', '1.0', '0.5', 'Assembly'])
        ws.append(['', 'Bearing', '2.0', '1.0', 'Bearing Assembly'])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="BOM_Sample_Template.xlsx"'
        return response
