import io
import csv
import openpyxl
from decimal import Decimal, InvalidOperation
from django.db import transaction
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser
from django.db.models import Sum
from urllib3 import request
from django.db import transaction

from .models import PurchaseEntry, StockLedger, InventoryAdjustment, FinishedGoodsStock, ScrapStock
from .serializers import (
    PurchaseEntrySerializer, StockLedgerSerializer,
    InventoryAdjustmentSerializer, FinishedGoodsStockSerializer, ScrapStockSerializer
)
from apps.master.models import RawMaterial
from apps.master.serializers import RawMaterialListSerializer
from rest_framework.views import APIView
from rest_framework.response import Response
from django.db import transaction
from apps.production.models import ProductionOrder
from apps.production.serializers import ProductionOrderSerializer


class PurchaseListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        entries = PurchaseEntry.objects.select_related('raw_material', 'created_by').all()
        item_id = request.query_params.get('item')
        if item_id:
            entries = entries.filter(raw_material_id=item_id)
        return Response(PurchaseEntrySerializer(entries, many=True).data)

    def post(self, request):
        if request.user.role not in ['ADMIN', 'STORE_MANAGER']:
            return Response({'error': 'Permission denied'}, status=403)
        serializer = PurchaseEntrySerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=400)

        with transaction.atomic():
            qty = Decimal(str(request.data['quantity']))
            rate = Decimal(str(request.data['purchase_rate']))
            gst_percent = Decimal(str(request.data.get('gst_percent', 18)))
            
            # Calculate total before GST
            subtotal = qty * rate
            # Calculate GST amount
            gst_amount = subtotal * (gst_percent / 100)
            # Total amount including GST
            total = subtotal + gst_amount

            material = RawMaterial.objects.select_for_update().get(pk=request.data['raw_material'])

            # Calculate moving average cost (use rate before GST for inventory valuation)
            current_value = material.current_stock * material.moving_avg_cost
            new_value = qty * rate
            new_total_qty = material.current_stock + qty
            new_avg_cost = (current_value + new_value) / new_total_qty if new_total_qty > 0 else rate

            material.current_stock += qty
            material.moving_avg_cost = round(new_avg_cost, 4)
            material.save()

            purchase = PurchaseEntry.objects.create(
                purchase_date=request.data['purchase_date'],
                supplier_name=request.data['supplier_name'],
                raw_material=material,
                quantity=qty,
                purchase_rate=rate,
                gst_percent=gst_percent,
                total_amount=total,
                notes=request.data.get('notes', ''),
                created_by=request.user
            )

            StockLedger.objects.create(
                raw_material=material,
                transaction_type='PURCHASE',
                quantity=qty,
                rate=rate,
                value=total,
                balance_qty=material.current_stock,
                balance_value=material.current_stock * material.moving_avg_cost,
                reference_no=f"PUR-{purchase.id}",
                notes=f"Purchase from {purchase.supplier_name}",
                created_by=request.user
            )

        return Response(PurchaseEntrySerializer(purchase).data, status=201)

class PurchaseDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def put(self, request, pk):

        if request.user.role not in ['ADMIN', 'STORE_MANAGER']:
            return Response({'error': 'Permission denied'}, status=403)

        with transaction.atomic():

            purchase = PurchaseEntry.objects.select_for_update().get(pk=pk)

            old_material = purchase.raw_material
            old_qty = purchase.quantity

            # reverse old stock
            old_material.current_stock -= old_qty
            old_material.save()

            # new material
            new_material = RawMaterial.objects.select_for_update().get(
                pk=request.data['raw_material']
            )

            qty = Decimal(str(request.data['quantity']))
            rate = Decimal(str(request.data['purchase_rate']))

            new_material.current_stock += qty
            new_material.save()

            purchase.purchase_date = request.data['purchase_date']
            purchase.supplier_name = request.data['supplier_name']
            purchase.raw_material = new_material
            purchase.quantity = qty
            purchase.purchase_rate = rate
            purchase.total_amount = qty * rate
            purchase.notes = request.data.get('notes', '')

            purchase.save()

        return Response(PurchaseEntrySerializer(purchase).data)

    def delete(self, request, pk):

        if request.user.role not in ['ADMIN', 'STORE_MANAGER']:
            return Response({'error': 'Permission denied'}, status=403)

        with transaction.atomic():  # ✅ REQUIRED

            purchase = PurchaseEntry.objects.select_for_update().get(pk=pk)

            material = purchase.raw_material
            qty = purchase.quantity

            # reverse stock
            material.current_stock -= qty
            material.save()

            purchase.delete()

        return Response({'message': 'Purchase deleted'})
class StockView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        items = RawMaterial.objects.filter(status=True)
        return Response(RawMaterialListSerializer(items, many=True).data)


class StockLedgerView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, item_id):
        entries = StockLedger.objects.filter(raw_material_id=item_id).select_related('created_by')
        return Response(StockLedgerSerializer(entries, many=True).data)


class AdjustmentListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        adjustments = InventoryAdjustment.objects.select_related('raw_material', 'created_by').all()
        return Response(InventoryAdjustmentSerializer(adjustments, many=True).data)

    def post(self, request):
        if request.user.role not in ['ADMIN', 'STORE_MANAGER']:
            return Response({'error': 'Permission denied'}, status=403)
        serializer = InventoryAdjustmentSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=400)

        with transaction.atomic():
            material = RawMaterial.objects.select_for_update().get(pk=request.data['raw_material'])
            qty = Decimal(str(request.data['quantity']))
            adj_type = request.data['adjustment_type']

            if adj_type == 'SUBTRACT' and material.current_stock < qty:
                return Response({'error': 'Insufficient stock for subtraction'}, status=400)

            if adj_type == 'ADD':
                material.current_stock += qty
                tx_type = 'ADJUSTMENT_ADD'
                sign_qty = qty
            else:
                material.current_stock -= qty
                tx_type = 'ADJUSTMENT_SUB'
                sign_qty = -qty

            material.save()

            adjustment = InventoryAdjustment.objects.create(
                raw_material=material,
                adjustment_type=adj_type,
                quantity=qty,
                reason=request.data['reason'],
                created_by=request.user
            )

            StockLedger.objects.create(
                raw_material=material,
                transaction_type=tx_type,
                quantity=sign_qty,
                rate=material.moving_avg_cost,
                value=sign_qty * material.moving_avg_cost,
                balance_qty=material.current_stock,
                balance_value=material.current_stock * material.moving_avg_cost,
                reference_no=f"ADJ-{adjustment.id}",
                notes=request.data['reason'],
                created_by=request.user
            )

        return Response(InventoryAdjustmentSerializer(adjustment).data, status=201)


class FinishedGoodsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Aggregate FG stock by product model
        from apps.master.models import ProductModel
        result = []
        for model in ProductModel.objects.filter(status=True):
            total_fg = FinishedGoodsStock.objects.filter(product_model=model).aggregate(
                total=Sum('quantity'))['total'] or 0
            total_scrap = ScrapStock.objects.filter(product_model=model).aggregate(
                total=Sum('quantity'))['total'] or 0
            result.append({
                'product_model': model.id,
                'model_id': model.model_id,
                'model_name': model.model_name,
                'finished_goods': float(total_fg),
                'scrap': float(total_scrap),
            })
        return Response(result)




class ProductionOrderDetailView(APIView):

    def put(self, request, pk):

        if request.user.role not in ['ADMIN', 'PRODUCTION_MANAGER']:
            return Response({'error': 'Permission denied'}, status=403)

        try:
            order = ProductionOrder.objects.get(pk=pk)
        except ProductionOrder.DoesNotExist:
            return Response({'error': 'Order not found'}, status=404)

        serializer = ProductionOrderSerializer(order, data=request.data, partial=True)

        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)

        return Response(serializer.errors, status=400)


    def delete(self, request, pk):

        if request.user.role not in ['ADMIN', 'PRODUCTION_MANAGER']:
            return Response({'error': 'Permission denied'}, status=403)

        try:
            order = ProductionOrder.objects.get(pk=pk)
        except ProductionOrder.DoesNotExist:
            return Response({'error': 'Order not found'}, status=404)

        order.delete()

        return Response({'message': 'Production order deleted'})


class ScrapStockView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        scrap = ScrapStock.objects.select_related('product_model').all()[:200]
        return Response(ScrapStockSerializer(scrap, many=True).data)


class ReorderAlertView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        items = RawMaterial.objects.filter(status=True, reorder_level__gt=0)
        alerts = [i for i in items if i.current_stock <= i.reorder_level]
        return Response(RawMaterialListSerializer(alerts, many=True).data)



class BulkOpeningStockImportView(APIView):
    """
    Import raw materials + opening stock from Excel or CSV.
    Columns: item_id, item_name, category, unit, opening_qty, unit_cost, reorder_level, lead_time
    - If item_id matches an existing material with stock=0 → update stock
    - If item_id matches an existing material with stock>0 → skip (already live)
    - If item_id not found → create new material + set opening stock
    """
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    VALID_CATEGORIES = {'ELECTRICAL', 'MECHANICAL', 'HARDWARE', 'PACKAGING', 'CONSUMABLE', 'OTHER'}
    VALID_UNITS = {'PCS', 'KG', 'MTR', 'LTR', 'SET', 'ROLL', 'BOX', 'GM', 'MM'}

    def _parse_file(self, file_obj):
        filename = file_obj.name.lower()
        if filename.endswith('.csv'):
            content = file_obj.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(content))
            return list(reader)
        elif filename.endswith(('.xlsx', '.xls')):
            wb = openpyxl.load_workbook(file_obj)
            ws = wb.active
            headers = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows())]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(v is not None and str(v).strip() for v in row):
                    rows.append(dict(zip(headers, [str(v).strip() if v is not None else '' for v in row])))
            return rows
        else:
            raise ValueError('Only .csv, .xlsx, .xls files are supported')

    def post(self, request):
        if request.user.role not in ['ADMIN', 'STORE_MANAGER']:
            return Response({'error': 'Permission denied'}, status=403)

        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'No file provided'}, status=400)

        try:
            rows = self._parse_file(file_obj)
        except Exception as e:
            return Response({'error': f'File parse error: {str(e)}'}, status=400)

        if not rows:
            return Response({'error': 'File is empty or has no data rows'}, status=400)

        created_items, updated_items, skipped_items, error_items = [], [], [], []

        for i, row in enumerate(rows, start=2):

            try:
                with transaction.atomic():

                    item_id = str(row.get('item_id', '')).strip()[:100]
                    item_name = str(row.get('item_name', '')).strip()[:300]

                    opening_qty = Decimal(row.get('opening_qty') or 0)
                    unit_cost = Decimal(row.get('unit_cost') or 0)
                    reorder_level = Decimal(row.get('reorder_level') or 0)
                    lead_time = int(row.get('lead_time') or 0)

                    category = str(row.get('category', 'OTHER')).upper()[:100]
                    unit = str(row.get('unit', 'PCS')).upper()[:20]

                    material = RawMaterial.objects.filter(item_id__iexact=item_id).first()

                    if material:

                        if material.current_stock > 0:
                            skipped_items.append(f"{material.item_name} already has stock")
                            continue

                        material.current_stock = opening_qty
                        material.moving_avg_cost = unit_cost
                        material.default_cost = unit_cost
                        material.reorder_level = reorder_level
                        material.lead_time = lead_time
                        material.save()

                        updated_items.append(material.item_name)

                    else:

                        material = RawMaterial.objects.create(
                            item_id=item_id,
                            item_name=item_name or item_id,
                            category=category,
                            unit=unit,
                            current_stock=opening_qty,
                            moving_avg_cost=unit_cost,
                            default_cost=unit_cost,
                            reorder_level=reorder_level,
                            lead_time=lead_time,
                            status=True
                        )

                        created_items.append(material.item_name)

                    if opening_qty > 0:
                        StockLedger.objects.create(
                            raw_material=material,
                            transaction_type='OPENING',
                            quantity=opening_qty,
                            rate=unit_cost,
                            value=opening_qty * unit_cost,
                            balance_qty=opening_qty,
                            balance_value=opening_qty * unit_cost,
                            reference_no='OPENING-IMPORT',
                            notes='Opening stock import',
                            created_by=request.user
                        )

            except Exception as e:
                error_items.append(f"Row {i}: {str(e)}")
                
        return Response({
            'created': created_items,
            'updated': updated_items,
            'skipped': skipped_items,
            'errors': error_items
        })

class OpeningStockSampleView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Opening Stock"

        headers = ['item_id', 'item_name', 'category', 'unit', 'opening_qty', 'unit_cost', 'reorder_level', 'lead_time']
        ws.append(headers)

        # Style header row
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='EA580C', end_color='EA580C', fill_type='solid')
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Sample rows (RO Booster Pump typical materials)
        sample_rows = [
            ['RM001', 'Motor Winding Wire 28 AWG', 'ELECTRICAL', 'KG', 10.5, 250.00, 2, 7],
            ['RM002', 'Pump Housing 100 GPD', 'MECHANICAL', 'PCS', 50, 180.00, 10, 14],
            ['RM003', 'Capacitor 10uF 250V', 'ELECTRICAL', 'PCS', 200, 15.50, 50, 5],
            ['RM004', 'O-Ring 50mm', 'HARDWARE', 'PCS', 500, 3.25, 100, 3],
            ['RM005', 'Shrink Sleeve 60mm', 'PACKAGING', 'PCS', 1000, 0.85, 200, 7],
            ['RM006', 'Bearing 6201', 'MECHANICAL', 'PCS', 150, 45.00, 30, 10],
        ]
        for row in sample_rows:
            ws.append(row)

        # Add a notes row
        ws.append([])
        notes_cell = ws.cell(row=len(sample_rows) + 3, column=1,
                             value='NOTES: category must be one of: ELECTRICAL, MECHANICAL, HARDWARE, PACKAGING, CONSUMABLE, OTHER')
        notes_cell.font = Font(italic=True, color='666666')
        ws.merge_cells(f'A{len(sample_rows) + 3}:H{len(sample_rows) + 3}')

        unit_cell = ws.cell(row=len(sample_rows) + 4, column=1,
                            value='unit must be one of: PCS, KG, MTR, LTR, SET, ROLL, BOX, GM, MM')
        unit_cell.font = Font(italic=True, color='666666')
        ws.merge_cells(f'A{len(sample_rows) + 4}:H{len(sample_rows) + 4}')

        # Column widths
        col_widths = [12, 30, 15, 8, 12, 12, 15, 12]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="Opening_Stock_Template.xlsx"'
        return response
